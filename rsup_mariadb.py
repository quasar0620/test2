import os
import re
import shutil
from openpyxl import load_workbook

def process_data(file_content):
    # Extract values between [RESULT_COMMENT] and [/RESULT_COMMENT]
    result_comment_match = re.search(r'\[RESULT\](.*?)\[/RESULT\]', file_content, re.DOTALL)
    result_comment_value = result_comment_match.group(1).strip() if result_comment_match else ''

    # Extract values between [DATA] and [/DATA]
    data_match = re.search(r'\[DATA\](.*?)\[/DATA\]', file_content, re.DOTALL)
    data_value = data_match.group(1).strip() if data_match else ''

    return f"{result_comment_value}\n\n[현황]\n{data_value}"

def copy_data_to_excel(data_path, template_path, result_path):
    # Copy the template file to the result path
    shutil.copyfile(template_path, result_path)
    result_workbook = load_workbook(result_path)

    file_names = [f for f in os.listdir(data_path) if f.endswith('.txt')]

    start_row = 5
    column_index = 3

    # Step 1: Copy data to '점검대상' sheet
    for file_name in file_names:
        file_path = os.path.join(data_path, file_name)
        with open(file_path, 'r', encoding='utf-8') as file:
            file_content = file.read()
            match = re.search(r'\[HOSTNAME\](.*?)\[/HOSTNAME\]', file_content, re.DOTALL)
            if match:
                data_content = match.group(1).strip()
            else:
                data_content = ""

        sheet = result_workbook.worksheets[2]
        sheet.cell(row=start_row, column=column_index, value=data_content)
        start_row += 1

    # Step 2: Copy 'SAMPLE' sheet for matching names
    for row in result_workbook.worksheets[2].iter_rows(min_row=5, min_col=column_index, max_col=column_index):
        for cell in row:
            if cell.value:
                sheet_copy = result_workbook.copy_worksheet(result_workbook['SAMPLE'])
                sheet_copy.title = cell.value

    # Step 3: Copy [SU-01] to [SU-60] data to corresponding sheets
    for i in range(1, 20):
        u_value = f'MY-{str(i).zfill(2)}'
        for file_name in file_names:
            file_path = os.path.join(data_path, file_name)
            with open(file_path, 'r', encoding='utf-8') as file:
                file_content = file.read()
                match_hostname = re.search(rf'\[HOSTNAME\](.*?)\[/HOSTNAME\]', file_content, re.DOTALL)
                match_u_value = re.search(rf'\[{u_value}\](.*?)\[/{u_value}\]', file_content, re.DOTALL)

                if match_hostname and match_u_value:
                    hostname = match_hostname.group(1).strip()
                    u_content = match_u_value.group(1).strip()

                    if hostname in result_workbook.sheetnames:
                        sheet = result_workbook[hostname]
                        for row in sheet.iter_rows(min_row=5, min_col=column_index):
                            for cell in row:
                                if cell.value == u_value:
                                    processed_content = process_data(u_content)
                                    sheet.cell(row=cell.row, column=cell.column + 6, value=processed_content)

    # Save the result workbook
    result_workbook.save(result_path)

def main():
    # 변경된 실행 위치 및 데이터 파일 경로
    dev_path = 'C:\\rsup\\dev'
    data_path = 'C:\\rsup\\data\\mariadb'
    template_path = os.path.join('C:\\rsup\\templet', 'templet_mariadb.xlsx')
    result_path = os.path.join('C:\\rsup', 'mariadb.xlsx')

    copy_data_to_excel(data_path, template_path, result_path)
    print(f"데이터를 엑셀에 성공적으로 삽입하고, {result_path}에 저장하였습니다.")

if __name__ == "__main__":
    main()
